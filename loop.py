#!/usr/bin/python
import os
import re
import sys
import time
import threading
import math
import pandas as pd
from openpyxl.worksheet import worksheet
import warnings
import re
from functools import reduce
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
# from xlsxwriter.workbook import Workbook
import matplotlib.pyplot as plt

LOCALS_NAME_IOS = [

    'Alana Hipermercado',
    'Atacadao',
    'Cestao da Economia',
	'Itao Supermercado',
	'Jaciana Supermercado',
	'Gbarbosa',
	'Frutaria e Mercadinho Claudinete',
	'Nenem Supermercados',
	'Supermercado Meira - Malhado',
	'Supermercado Meira - Centro',
	'Supermercado Meira - N. Senhora da Vitoria',
	'Supermercado Meira - Vilela',
	'Supermercado Meira - N. Costa',
	'Supermercado Mangostao',

]

LOCALS_IOS = [
    'Alana Hipermercado',
    'Atacadao',
    'NAO POSSUI',
    'Supermercados Itao',
   	'Mercado Jaciana',
   	'GBARBOSA',
   	'Frutaria e Mercadinho Claudinete',
   	'Meu Mercado',
   	'Supermercado Dalnorde',
   	'Dalnorde Supermercados',
   	'Dalnorde Supermercados',
   	'Supermercado Meira',
   	'Dalnorde Supermercado',
   	'Supermercado Mangostao LTDA',
]

LOCALS_NAME_ITN = [

    'Atacadao Rondelli',
    'Bom de preço',
    'Bom Preco',
	'Compre Aqui - Canal',
	'Compre Aqui',
	'Mercado Mattos',
	'Mercado Dois Imaos',
	'Maxx Atacado',
	'Novo Barateiro',
	'Supermercado Meira',
	'Supermercado Barateiro',
	'Supermercado Itao - Centro',
	'Supermercado Itao - S.C',

]

LOCALS_ITN = [
    'RONDELLI COMERCIO E TRANSPORTE LTDA',
    'BOM DE PRECO',
    'BOMPRECO',
	'SUPERMERCADOS COMPRE AQUI',
	'SUPERMERCADOS COMPRE AQUI',
	'MERCADO MATOS',
	'MERCADO IRMÃOS',
	'MAXXI',
	'NOVO BARATEIRO',
	'SUPERMERCADOS DALNORDES',
	'ITAO',
	'ITAO SUPERMERCADOS IMPORTACOES E EXPORTACOES S/A',
	'HIPER ITAO',
]


LOCALS = {

	'Itabuna' : {

		'file_name' : LOCALS_NAME_ITN,
		'estab_name': LOCALS_ITN

	},
	'Ilhéus' : {

		'file_name' : LOCALS_NAME_IOS,
		'estab_name': LOCALS_IOS

	}

}

PRODUCTS = {'ACUCAR CRISTAL': ['ACUCAR CRISTAL', 'ACUCAR CRISTAL 1KG'],
            'ARROZ PARBOILIZADO': ['ARROZ PARBOILIZADO', 'ARROZ PARBOILIZADO 1KG'],
            'BANANA PRATA': ['BANANA DA PRATA', 'BANANA PRATA', 'BANANA KG'],
            'CAFE MOIDO': ['CAFE 250G', 'CAFE MOIDO'],
            'CHA DENTRO': ['CHA DE DENTRO', 'COXAO MOLE', 'CARNE BOVINA CHA DE DENTRO'],
            'FARINHA DE MANDIOCA': ['FARINHA DE MANDIOCA', 'FARINHA MAND', 'FARINHA MANDIOCA'],
            'FEIJAO CARIOCA': ['FEIJAO CARIOCA'],
            'LEITE LIQUIDO': ['LEITE LIQUIDO'],
            'MANTEIGA 500G': ['MANTEIGA 500G', 'MANTEIGA'],
            'OLEO SOJA': ['OLEO DE SOJA', 'OLEO 900ML'],
            'PAO FRANCES': ['PAO FRANCES', 'PAO FRANCES KG'],
            'TOMATE KG': ['TOMATE KG']}

FILTERED_PATH = 'Coleta por Data'
UNIFIED_PATH = 'Coletas Concatenadas'
TODOS = 'Coleta por Arquivo'
NORMAL = 'Coleta Padrao'
EXTRA = 'Coleta Extra'

class Filter_Colect_Todos:

	def __init__(self,  city, products, city_info, tk=None, change_frame=None):

		self.city = {"Itabuna": {'folders': []}, "Ilhéus": {"folders": []}}
		self.estab = []
		if city != None:
			for city_name in city:

				self.city[city_name] = {'folders': []}

				for index, (estab, adress, estab_name) in enumerate(city_info[city_name]):

					self.estab.append({'name': estab_name.upper(), 'adress': adress.upper(), 'file_name': estab})

		self.tk = tk
		self.change_frame = change_frame
		self.products = products
		self.city_info = city_info
		self.path = Path().absolute()
		self.thread_right = []
		self.thread = []

	def unify_ind_analisys(self):

		with os.scandir(UNIFIED_PATH) as directory:

			global_append = []
			for file in directory:

				if file.is_file() and file.name.endswith('.xlsx'):

					graph_path = '{}/{} Gráficos'.format(UNIFIED_PATH, file.name.replace('.xlsx', ''))
					file_path = '{}/{}'.format(UNIFIED_PATH, file.name)
					if not os.path.exists(graph_path):

						os.makedirs(graph_path)

					print(file_path)
					df = pd.read_excel(file_path, skiprows=0, index_col=0)

					df['PRECO'] = df['PRECO'].map(lambda x: x if len(x) <= 9 else 'R$ 0.00')
					df['PRECO'] = df['PRECO'].map(lambda x: x.lstrip('R$ ').rstrip('')).astype(float)
					append_df = []
					for product_name, products in PRODUCTS.items():

						product_df = []
						for product in products:

							# product = product.split(' ')
							new_df = ''
							# new_df = df[df.apply(lambda r: any([kw in r[0] for kw in product]), axis=1)]
							new_df = df[df.KEYWORD == product]
							new_df = new_df.rename(columns={'PRECO': product_name})
							new_df = new_df[product_name].sort_values()
							product_df.append(new_df)

						product_df = pd.concat(product_df)
						product_df = product_df.drop_duplicates(keep="first")
						product_df = product_df.reset_index()
						# print(product_df)
						append_df.append(product_df)

					named_df = pd.concat(append_df, axis=1)
					named_df = named_df.drop(columns='index',axis=1)
					global_append.append(named_df)
					# ax = named_df.plot(kind='barh', title=file.name.replace(
						# '.xlsx', ''), legend=True, figsize=(10, 10), fontsize=10, width=1, align='center')
					# ax.set_xlabel("Preços", fontsize=10)
					# ax.set_ylabel("Preços", fontsize=10)
					# ax.get_figure()

			global_df = pd.concat(global_append)
			print(global_df)
			ax = global_df.plot(kind='bar', title=file.name.replace(
							'.xlsx', ''), legend=True, figsize=(10, 10), fontsize=10, width=1, align='center')
			ax.set_xlabel("Preços", fontsize=10)
			ax.set_ylabel("Preços", fontsize=10)
			ax.get_figure()
			plt.show()
			sys.exit(1)

		return 0

	def unify_colect(self, type=TODOS):

		for city_name in self.city:

			path = "{}\{}\{}".format(FILTERED_PATH, city_name, type)
			uni_path = "{}\{}\{}".format(UNIFIED_PATH, city_name, type)

			with os.scandir(path) as directory:

				for file in directory:

					if file.is_file() and file.name.endswith('.xlsx'):

						print('UNIFICANDO ARQUIVOS ... {}\t  Cidade :{}'.format(file.name, city_name))

						all_dfs = pd.read_excel('{}\{}'.format(path, file.name), sheet_name=None, index_col=0)

						df = pd.concat(all_dfs, ignore_index=True)
						writer = pd.ExcelWriter('{}\{}'.format(uni_path, file.name), engine='openpyxl')
						df = df.drop_duplicates(subset=["PRODUTO", "ESTABELECIMENTO", "ENDERECO","PRECO"], keep="first")

						df = df.sort_values(by=['KEYWORD', 'PRECO'])

						df.to_excel(writer, sheet_name='Coleta Geral',index=False, startrow=0, startcol=1)


						border = Border(left=Side(border_style='thin', color='FF000000'),
											right=Side(border_style='thin',
													color='FF000000'),
											top=Side(border_style='thin',
													color='FF000000'),
											bottom=Side(border_style='thin',
														color='FF000000'),
											diagonal=Side(border_style='thin', color='FF000000'), diagonal_direction=0,
											outline=Side(
												border_style='thin', color='FF000000'),
											vertical=Side(
												border_style='thin', color='FF000000'),
											horizontal=Side(border_style='thin', color='FF000000'))

						workbook = writer.book['Coleta Geral']
						worksheet = workbook
						for cell in worksheet['B']:

							cell.border = border
							cell.alignment = Alignment(horizontal='center')

						for cell in worksheet['C']:

							cell.border = border
							cell.alignment = Alignment(horizontal='center')

						for cell in worksheet['D']:

							cell.border = border
							cell.alignment = Alignment(horizontal='center')

						for cell in worksheet['E']:

							cell.border = border
							cell.alignment = Alignment(horizontal='center')

						for cell in worksheet['F']:

							cell.border = border
							cell.alignment = Alignment(horizontal='center')

						for col in worksheet.columns:
							max_length = 0
							column = col[0].column_letter  # Get the column name
							for cell in col:
								try:  # Necessary to avoid error on empty cells
									if len(str(cell.value)) > max_length:
										max_length = len(str(cell.value))
								except:
									pass
							adjusted_width = (max_length + 2) * 1.2
							worksheet.column_dimensions[column].width = adjusted_width

						# sys.exit(1)

						writer.save()

	def filter_xlsx(self,file_name, city_name, date):

		print('FILTRANDO ARQUIVOS ... {}\t CIDADE : {}'.format(file_name, city_name))
		df = pd.read_excel(file_name, skiprows=0, index_col=0)
		estab_list = df.drop_duplicates(subset=["ESTABELECIMENTO"], keep="first")
		estab_adr_list = estab_list['ENDERECO']
		estab_list = estab_list['ESTABELECIMENTO']
		estab_list = list(filter(lambda estab: estab.isupper(), estab_list))
		estab_adr_list = list(filter(lambda estab: estab.isupper(), estab_adr_list))
		local = self.estab

		appended_data = []
		for (product, keyword) in self.products:

			keyword = keyword.split(' ')
			appended_data.append(
                            df[df.apply(lambda r: all([kw in r[0] for kw in keyword]), axis=1)])

		df = pd.concat(appended_data)
		df = df.drop_duplicates(
			subset=["PRODUTO", "ESTABELECIMENTO", "ENDERECO", "PRECO"], keep="first")
		df = df.sort_values(by=['KEYWORD', 'PRECO'])

		for i, estab in enumerate(estab_list):

			info = list(filter(lambda x: x['name'] == estab, local))
			if len(info) != 0:

				info = info[0]
				new_file = re.sub('[^A-Za-z0-9,]+', ' ', info['file_name'])
				adress = re.sub('[^A-Za-z0-9,]+', ' ', info['adress'])
				new_folder = "{}\{}\{}".format(FILTERED_PATH, city_name, TODOS)

			else:

				new_file = re.sub('[^A-Za-z0-9,]+', ' ', estab)
				adress = re.sub('[^A-Za-z0-9,]+', ' ', estab_adr_list[i])
				new_folder = "{}\{}\{}".format(FILTERED_PATH, city_name, EXTRA)

			temp_df = df
			path = "{}\{}.xlsx".format(new_folder, new_file)

			mask = temp_df.ENDERECO.str.contains(adress.upper())
			temp_df = temp_df[mask]

			if len(temp_df) != 0:

				try:

					file = load_workbook(path)
					writer = pd.ExcelWriter(path, engine='openpyxl')
					writer.book = file
					if date in file.sheetnames:
						count = file.sheetnames.count(date)
						date = "{} {}".format(date, count + 1)
				except:

					writer = pd.ExcelWriter(path, engine='openpyxl')

				temp_df = temp_df.to_excel(
					writer, sheet_name=date,  index=False, startrow=0, startcol=1)
				border = Border(left=Side(border_style='thin', color='FF000000'),
                                    right=Side(border_style='thin',
                                               color='FF000000'),
                                    top=Side(border_style='thin',
                                             color='FF000000'),
                                    bottom=Side(border_style='thin',
                                                color='FF000000'),
                                    diagonal=Side(border_style='thin', color='FF000000'), diagonal_direction=0,
                                    outline=Side(border_style='thin',
                                                 color='FF000000'),
                                    vertical=Side(border_style='thin',
                                                  color='FF000000'),
                                    horizontal=Side(border_style='thin', color='FF000000'))

				workbook = writer.book[date]
				worksheet = workbook
				for cell in worksheet['B']:

					cell.border = border

				for cell in worksheet['C']:

					cell.border = border

				for cell in worksheet['D']:

					cell.border = border

				for cell in worksheet['E']:

					cell.border = border

				for cell in worksheet['F']:

					cell.border = border

				for col in worksheet.columns:
					max_length = 0
					column = col[0].column_letter  # Get the column name
					for cell in col:
						try:  # Necessary to avoid error on empty cells
							if len(str(cell.value)) > max_length:
								max_length = len(str(cell.value))
						except:
							pass
					adjusted_width = (max_length + 2) * 1.2
					worksheet.column_dimensions[column].width = adjusted_width

				writer.save()

	def filter_folder(self,folders, city_name):

		for folder in folders:

			for file in folder['files']:

				# print(file['file_name'])
				# print(folder['date'])
				# print(city_name)
				# self.filter_xlsx("{}\{}".format(folder['folder_name'], file['file_name']), city_name, folder['date'])
				self.filter_xlsx_unique("{}\{}".format(folder['folder_name'], file['file_name']), city_name, folder['date'])

	def filter_xlsx_unique(self, file_name, city_name, date):

		print('FILTRANDO ARQUIVOS ... {}\t CIDADE : {}'.format(file_name, city_name))
		df = pd.read_excel(file_name, skiprows=0, index_col=0)
		estab_list = df.drop_duplicates(subset=["ESTABELECIMENTO"], keep="first")
		estab_adr_list = estab_list['ENDERECO']
		estab_list = estab_list['ESTABELECIMENTO']
		estab_list = list(filter(lambda estab: estab.isupper(), estab_list))
		estab_adr_list = list(filter(lambda estab: estab.isupper(), estab_adr_list))
		local = self.estab

		appended_data = []
		for (product, keyword) in self.products:

				keyword = keyword.split(' ')
				appended_data.append(df[df.apply(lambda r: all([kw in r[0] for kw in keyword]), axis=1)])

		df = pd.concat(appended_data)
		df = df.drop_duplicates(subset=["PRODUTO", "ESTABELECIMENTO", "ENDERECO", "PRECO"], keep="first")
		df = df.sort_values(by=['KEYWORD', 'PRECO'])

		for i,estab in enumerate(estab_list):

			info = list(filter(lambda x: x['name'].upper() == estab.upper(), local))
			if len(info) != 0:

				info = info[0]
				new_file = info['file_name']
				adress = re.sub('[^A-Za-z0-9,/-]+', ' ', info['adress'])
				new_folder = "{}\{}\{}".format(FILTERED_PATH, city_name, TODOS)

			else:

				new_file = re.sub('[^A-Za-z0-9,-]+', ' ', estab)
				adress = re.sub('[^A-Za-z0-9,/-]+', ' ', estab_adr_list[i])
				new_folder = "{}\{}\{}".format(FILTERED_PATH, city_name, EXTRA)

			temp_df = df
			path = "{}\{}.xlsx".format(new_folder, new_file)

			temp_df = temp_df[temp_df.ESTABELECIMENTO.str.upper() == estab.upper()]
			temp_df = temp_df[temp_df.ENDERECO.str.contains(adress.upper())]

			if len(temp_df) != 0:

				try:

					file = load_workbook(path)
					writer = pd.ExcelWriter(path, engine='openpyxl')
					writer.book = file
					if date in file.sheetnames:
						count = file.sheetnames.count(date)
						date = "{} {}".format(date, count + 1)
				except:

					writer = pd.ExcelWriter(path, engine='openpyxl')

				temp_df = temp_df.to_excel(
					writer, sheet_name=date,  index=False, startrow=0, startcol=1)
				border = Border(left=Side(border_style='thin', color='FF000000'),
									right=Side(border_style='thin',
											color='FF000000'),
									top=Side(border_style='thin',
											color='FF000000'),
									bottom=Side(border_style='thin',
												color='FF000000'),
									diagonal=Side(border_style='thin', color='FF000000'), diagonal_direction=0,
									outline=Side(border_style='thin',
												color='FF000000'),
									vertical=Side(border_style='thin',
												color='FF000000'),
									horizontal=Side(border_style='thin', color='FF000000'))

				workbook = writer.book[date]
				worksheet = workbook
				for cell in worksheet['B']:

					cell.border = border
					cell.alignment = Alignment(horizontal='center')

				for cell in worksheet['C']:

					cell.border = border
					cell.alignment = Alignment(horizontal='center')

				for cell in worksheet['D']:

					cell.border = border
					cell.alignment = Alignment(horizontal='center')

				for cell in worksheet['E']:

					cell.border = border
					cell.alignment = Alignment(horizontal='center')

				for cell in worksheet['F']:

					cell.border = border
					cell.alignment = Alignment(horizontal='center')

				for col in worksheet.columns:
					max_length = 0
					column = col[0].column_letter  # Get the column name
					for cell in col:
						try:  # Necessary to avoid error on empty cells
							if len(str(cell.value)) > max_length:
								max_length = len(str(cell.value))
						except:
							pass
					adjusted_width = (max_length + 2) * 1.2
					worksheet.column_dimensions[column].width = adjusted_width

				writer.save()

	def filter_thread(self,city):

		for city_name in city:

			folder = city[city_name]
			folder = folder['folders']
			self.filter_folder(folder, city_name)
			# size = len(folder)
			# left = math.ceil(size/2)
			# right = left + 1
			# folder_left = folder[:left]
			# folder_right = folder[right:]
			# thread_left = threading.Thread(target=lambda : self.filter_folder(folder_left, city_name))
			# self.thread.append(thread_left)
			# thread_left.start()
			# thread_right = threading.Thread(target=lambda : self.filter_folder(folder_right, city_name))
			# self.thread.append(thread_right)
			# thread_right.start()

	def get_files(self,folder, city_name):

		prev_date = ''
		with os.scandir(folder['folder_name']) as directory:

			for file in directory:

				if file.is_file() and file.name.startswith('TODOS') and file.name.endswith('.xlsx'):

					self.filter_xlsx_unique("{}\{}".format(folder['folder_name'], file.name), city_name, folder['date'])
					# thread = threading.Thread(target=lambda : filter_xlsx("{}\{}".format(folder['folder_name'], file.name), city_name, date))
					# thread.start()
					folder['files'].append({'file_name': file.name})

		return folder

	def get_folders(self,city, path):

		with os.scandir(path) as directory:

			folders = []
			for folder in directory:

				for city_name in city:

					if folder.is_dir() and folder.name.startswith(city_name):

						date = folder.name.replace(city_name, '')
						date = date.replace('[', '')
						date = date.replace(']', '')
						date = date.replace(' ', '')
						city[city_name]['folders'].append(
							self.get_files({"folder_name": folder.name, "date": date, 'files': []}, city_name))

		return city

	def create_dir(self):

		if not os.path.exists(FILTERED_PATH):

			os.makedirs(FILTERED_PATH)

		if not os.path.exists(UNIFIED_PATH):

			os.makedirs(UNIFIED_PATH)

		for city_name in self.city:

			new_folder = "{}\{}\{}".format(FILTERED_PATH, city_name, NORMAL)

			if not os.path.exists(new_folder):

				os.makedirs(new_folder)

			new_folder = "{}\{}\{}".format(FILTERED_PATH, city_name, EXTRA)

			if not os.path.exists(new_folder):

				os.makedirs(new_folder)

			new_folder = "{}\{}\{}".format(FILTERED_PATH, city_name, TODOS)

			if not os.path.exists(new_folder):

				os.makedirs(new_folder)

			new_folder = "{}\{}\{}".format(UNIFIED_PATH, city_name, NORMAL)

			if not os.path.exists(new_folder):

				os.makedirs(new_folder)

			new_folder = "{}\{}\{}".format(UNIFIED_PATH, city_name, EXTRA)

			if not os.path.exists(new_folder):

				os.makedirs(new_folder)

			new_folder = "{}\{}\{}".format(UNIFIED_PATH, city_name, TODOS)

			if not os.path.exists(new_folder):

				os.makedirs(new_folder)

	def start(self):

		self.create_dir()
		self.city = self.get_folders(self.city, self.path)
		self.filter_thread(self.city)

		for thread in self.thread:

				thread.join()

		self.unify_colect(TODOS)
		self.unify_colect(EXTRA)

class Filter_Colect:

	def __init__(self, city, tk=None, change_frame=None):

		self.city = {"Itabuna": {'folders': []}, "Ilhéus": {"folders": []}}
		if city != None:
			for city_name in city:
				self.city[city_name] = {'folders': []}

		self.tk = tk
		self.change_frame = change_frame
		self.path = Path().absolute()
		self.thread_right = []
		self.thread = []

	def unify_ind_analisys(self):

		with os.scandir(UNIFIED_PATH) as directory:

			global_append = []
			for file in directory:

				if file.is_file() and file.name.endswith('.xlsx'):

					graph_path = '{}/{} Gráficos'.format(UNIFIED_PATH, file.name.replace('.xlsx', ''))
					file_path = '{}/{}'.format(UNIFIED_PATH, file.name)
					if not os.path.exists(graph_path):

						os.makedirs(graph_path)

					print(file_path)
					df = pd.read_excel(file_path, skiprows=0, index_col=0)

					df['PRECO'] = df['PRECO'].map(lambda x: x if len(x) <= 9 else 'R$ 0.00')
					df['PRECO'] = df['PRECO'].map(lambda x: x.lstrip('R$ ').rstrip('')).astype(float)
					append_df = []
					for product_name, products in PRODUCTS.items():

						product_df = []
						for product in products:

							# product = product.split(' ')
							new_df = ''
							# new_df = df[df.apply(lambda r: any([kw in r[0] for kw in product]), axis=1)]
							new_df = df[df.KEYWORD == product]
							new_df = new_df.rename(columns={'PRECO': product_name})
							new_df = new_df[product_name].sort_values()
							product_df.append(new_df)

						product_df = pd.concat(product_df)
						product_df = product_df.drop_duplicates(keep="first")
						product_df = product_df.reset_index()
						# print(product_df)
						append_df.append(product_df)

					named_df = pd.concat(append_df, axis=1)
					named_df = named_df.drop(columns='index',axis=1)
					global_append.append(named_df)
					# ax = named_df.plot(kind='barh', title=file.name.replace(
						# '.xlsx', ''), legend=True, figsize=(10, 10), fontsize=10, width=1, align='center')
					# ax.set_xlabel("Preços", fontsize=10)
					# ax.set_ylabel("Preços", fontsize=10)
					# ax.get_figure()

			global_df = pd.concat(global_append)
			print(global_df)
			ax = global_df.plot(kind='bar', title=file.name.replace(
							'.xlsx', ''), legend=True, figsize=(10, 10), fontsize=10, width=1, align='center')
			ax.set_xlabel("Preços", fontsize=10)
			ax.set_ylabel("Preços", fontsize=10)
			ax.get_figure()
			plt.show()


			sys.exit(1)

		return 0

	def unify_colect(self):

		for city_name in self.city:

			path = "{}\{}\{}".format(FILTERED_PATH, city_name, NORMAL)
			uni_path = "{}\{}\{}".format(UNIFIED_PATH, city_name, NORMAL)

			with os.scandir(path) as directory:

				for file in directory:

					if file.is_file() and file.name.endswith('.xlsx'):

						all_dfs = pd.read_excel('{}\{}'.format(path, file.name), sheet_name=None, index_col=0)

						print("UNIFICANDO ARQUIVOS ... {}\t Cidade: {}".format(file.name, city_name))
						df = pd.concat(all_dfs, ignore_index=True)
						writer = pd.ExcelWriter('{}\{}'.format(uni_path, file.name), engine='openpyxl')
						df = df.drop_duplicates(subset=["PRODUTO", "ESTABELECIMENTO", "PRECO"], keep="first")

						df = df.sort_values(by=['KEYWORD'])

						df.to_excel(writer, sheet_name='Coleta Geral',index=False, startrow=0, startcol=1)


						border = Border(left=Side(border_style='thin', color='FF000000'),
											right=Side(border_style='thin',
													color='FF000000'),
											top=Side(border_style='thin',
													color='FF000000'),
											bottom=Side(border_style='thin',
														color='FF000000'),
											diagonal=Side(border_style='thin', color='FF000000'), diagonal_direction=0,
											outline=Side(
												border_style='thin', color='FF000000'),
											vertical=Side(
												border_style='thin', color='FF000000'),
											horizontal=Side(border_style='thin', color='FF000000'))

						workbook = writer.book['Coleta Geral']
						worksheet = workbook
						for cell in worksheet['B']:

							cell.border = border
							cell.alignment = Alignment(horizontal='center')

						for cell in worksheet['C']:

							cell.border = border
							cell.alignment = Alignment(horizontal='center')

						for cell in worksheet['D']:

							cell.border = border
							cell.alignment = Alignment(horizontal='center')

						for cell in worksheet['E']:

							cell.border = border
							cell.alignment = Alignment(horizontal='center')

						for cell in worksheet['F']:

							cell.border = border
							cell.alignment = Alignment(horizontal='center')

						for col in worksheet.columns:
							max_length = 0
							column = col[0].column_letter  # Get the column name
							for cell in col:
								try:  # Necessary to avoid error on empty cells
									if len(str(cell.value)) > max_length:
										max_length = len(str(cell.value))
								except:
									pass
							adjusted_width = (max_length + 2) * 1.2
							worksheet.column_dimensions[column].width = adjusted_width

						# sys.exit(1)

						writer.save()

	def fix_key_end(self, file_name, city_name, date, _file_name):


		df = pd.read_excel(file_name, skiprows=0, index_col=0)

		try:

			print(df['ENDEREÇO'])

		except:

			return

		# print('LENDO ARQUIVOS ... {} , CIDADE : {}'.format(file_name, city_name))
		new_folder = "{}\{}\{}".format(FILTERED_PATH, city_name, NORMAL)
		estab_list = df.drop_duplicates(subset=["ESTABELECIMENTO"], keep="first")
		estab_list = estab_list['ESTABELECIMENTO']
		estab_list = list(filter(lambda estab: estab.isupper(), estab_list))

		path = "{}\{}.xlsx".format(new_folder, _file_name)

		df = df.reset_index(drop=True)
		key = list(PRODUCTS.values())
		key = reduce(lambda a, b: a+b, key)

		for i, row in df.iterrows():

			coluna = row['KEYWORD']
			estab = row['ENDEREÇO']
			if any([x for x in key if x in estab]):

				df.at[i, 'KEYWORD'] = estab
				df.at[i, 'ENDEREÇO'] = coluna

		writer = pd.ExcelWriter(path, engine='openpyxl')

		df = df.to_excel(
			writer, sheet_name=date,  index=False, startrow=0, startcol=1)
		border = Border(left=Side(border_style='thin', color='FF000000'),
						right=Side(border_style='thin', color='FF000000'),
						top=Side(border_style='thin',
								color='FF000000'),
						bottom=Side(border_style='thin',
									color='FF000000'),
						diagonal=Side(border_style='thin', color='FF000000'), diagonal_direction=0,
						outline=Side(border_style='thin',
									color='FF000000'),
						vertical=Side(border_style='thin',
										color='FF000000'),
						horizontal=Side(border_style='thin', color='FF000000'))

		workbook = writer.book[date]
		worksheet = workbook
		for cell in worksheet['B']:

				cell.border = border

		for cell in worksheet['C']:

			cell.border = border

		for cell in worksheet['D']:

			cell.border = border

		for cell in worksheet['E']:

			cell.border = border

		for cell in worksheet['F']:

			cell.border = border

		for col in worksheet.columns:
			max_length = 0
			column = col[0].column_letter  # Get the column name
			for cell in col:
				try:  # Necessary to avoid error on empty cells
					if len(str(cell.value)) > max_length:
						max_length = len(str(cell.value))
				except:
					pass
			adjusted_width = (max_length + 2) * 1.2
			worksheet.column_dimensions[column].width = adjusted_width

		writer.save()

	def filter_xlsx_unique(self, file_name, city_name, date, _file_name):

		print('FILTRANDO ARQUIVOS ... {}\t CIDADE : {}'.format(file_name, city_name))
		new_folder = "{}\{}\{}".format(FILTERED_PATH, city_name, NORMAL)
		df = pd.read_excel(file_name, skiprows=0, index_col=0)

		new_file = _file_name.replace('.xlsx', '')
		temp_df = df
		path = "{}\{}.xlsx".format(new_folder, new_file)

		try:

			file = load_workbook(path)
			writer = pd.ExcelWriter(path, engine='openpyxl')
			writer.book = file
			if date in file.sheetnames:
				count = file.sheetnames.count(date)
				date = "{} {}".format(date, count + 1)

		except:

			writer = pd.ExcelWriter(path, engine='openpyxl')

		temp_df = temp_df.to_excel(
			writer, sheet_name=date,  index=False, startrow=0, startcol=1)
		border = Border(left=Side(border_style='thin', color='FF000000'),
						right=Side(border_style='thin', color='FF000000'),
						top=Side(border_style='thin', color='FF000000'),
						bottom=Side(border_style='thin', color='FF000000'),
						diagonal=Side(border_style='thin', color='FF000000'), diagonal_direction=0,
						outline=Side(border_style='thin',
									color='FF000000'),
						vertical=Side(border_style='thin',
									color='FF000000'),
						horizontal=Side(border_style='thin', color='FF000000'))

		workbook = writer.book[date]
		worksheet = workbook
		for cell in worksheet['B']:

			cell.border = border
			cell.alignment = Alignment(horizontal='center')

		for cell in worksheet['C']:

			cell.border = border
			cell.alignment = Alignment(horizontal='center')

		for cell in worksheet['D']:

			cell.border = border
			cell.alignment = Alignment(horizontal='center')

		for cell in worksheet['E']:

			cell.border = border
			cell.alignment = Alignment(horizontal='center')

		for cell in worksheet['F']:

			cell.border = border
			cell.alignment = Alignment(horizontal='center')

		for col in worksheet.columns:
			max_length = 0
			column = col[0].column_letter  # Get the column name
			for cell in col:
				try:  # Necessary to avoid error on empty cells
					if len(str(cell.value)) > max_length:
						max_length = len(str(cell.value))
				except:
					pass
			adjusted_width = (max_length + 2) * 1.2
			worksheet.column_dimensions[column].width = adjusted_width

		writer.save()

			# print("\n\n ------------------------------------------ {} -----------------------------------------".format(estab.upper()))

	def filter_folder(self,folders, city_name):

		for folder in folders:

			for file in folder['files']:

				self.filter_xlsx("{}\{}".format(folder['folder_name'], file['file_name']), city_name, folder['date'], file['file_name'])

	def filter_thread(self,city):

		for city_name in city:

			folder = city[city_name]
			folder = folder['folders']
			size = len(folder)
			left = math.ceil(size/2)
			right = left + 1
			folder_left = folder[:left]
			folder_right = folder[right:]
			thread_left = threading.Thread(target=lambda : self.filter_folder(folder_left, city_name))
			self.thread.append(thread_left)
			thread_left.start()
			thread_right = threading.Thread(target=lambda : self.filter_folder(folder_right, city_name))
			self.thread.append(thread_right)
			thread_right.start()

	def get_files(self,folder, city_name):

		with os.scandir(folder['folder_name']) as directory:

			for file in directory:

				if file.is_file() and not file.name.startswith('TODOS') and file.name.endswith('.xlsx'):

					self.filter_xlsx_unique("{}\{}".format(folder['folder_name'], file.name), city_name, folder['date'], file.name)
					# self.fix_key_end("{}\{}".format(folder['folder_name'], file.name), city_name, folder['date'], file.name)
					folder['files'].append({'file_name': file.name})


		return folder

	def get_folders(self,city, path):

		with os.scandir(path) as directory:

			folders = []

			for folder in directory:

				for city_name in city:

					if folder.is_dir() and folder.name.startswith(city_name):

						date = folder.name.replace(city_name, '')
						date = date.replace('[', '')
						date = date.replace(']', '')
						date = date.replace(' ', '')
						city[city_name]['folders'].append(
							self.get_files({"folder_name": folder.name, "date": date, 'files': []}, city_name))

		return city

	def create_dir(self):

		if not os.path.exists(FILTERED_PATH):

			os.makedirs(FILTERED_PATH)

		if not os.path.exists(UNIFIED_PATH):

			os.makedirs(UNIFIED_PATH)

		for city_name in self.city:

			new_folder = "{}\{}\{}".format(FILTERED_PATH, city_name, NORMAL)

			if not os.path.exists(new_folder):

				os.makedirs(new_folder)

			new_folder = "{}\{}\{}".format(FILTERED_PATH, city_name, EXTRA)

			if not os.path.exists(new_folder):

				os.makedirs(new_folder)

			new_folder = "{}\{}\{}".format(FILTERED_PATH, city_name, TODOS)

			if not os.path.exists(new_folder):

				os.makedirs(new_folder)

			new_folder = "{}\{}\{}".format(UNIFIED_PATH, city_name, NORMAL)

			if not os.path.exists(new_folder):

				os.makedirs(new_folder)

			new_folder = "{}\{}\{}".format(UNIFIED_PATH, city_name, EXTRA)

			if not os.path.exists(new_folder):

				os.makedirs(new_folder)

			new_folder = "{}\{}\{}".format(UNIFIED_PATH, city_name, TODOS)

			if not os.path.exists(new_folder):

				os.makedirs(new_folder)

	def start(self):

		self.create_dir()
		self.city = self.get_folders(self.city, self.path)
		self.unify_colect()
