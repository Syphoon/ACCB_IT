import os
import sys
import pandas as pd
import ctypes
from openpyxl.styles import Border, Side, Alignment
from openpyxl import load_workbook
import itertools

FILE_ATTRIBUTE_HIDDEN = 0x02


LOCALS_ADR_ITN = {

    'Atacadao Rondelli': {'adress': 'AVENIDA JOSE SOARES PINHEIRO 2700', 'web_name': 'RONDELLI COMERCIO E TRANSPORTE LTDA'},
    'Bom de Preco': {'adress': 'RUA BARAO DO RIO BRANCO 203', 'web_name': 'BOM DE PRECO'},
   	'Bom Preco': {'adress': 'AVENIDA AZIZ MARON S/N', 'web_name': 'BOMPRECO'},
   	'Compre Aqui - Canal': {'adress': 'PRACA DOS CAPUCHINHOS 238', 'web_name': 'SUPERMERCADOS COMPRE AQUI'},
   	'Compre Aqui': {'adress': 'AVENIDA AMELIA AMADO 358', 'web_name': 'SUPERMERCADOS COMPRE AQUI'},
   	'Mercado Mattos': {'adress': 'RUA C 390', 'web_name': 'MERCADO MATOS'},
   	'Mercado Dois Imaos': {'adress': 'AVENIDA BIONOR REBOUÇAS BRANDÃO 368', 'web_name': 'MERCADO IRMÃOS'},
	'Maxx Atacado': {'adress':'RODOVIA ITABUNA - ILHEUS, KM 2, BR 415 S/N', 'web_name': 'MAXXI'},
	'Novo Barateiro': {'adress': 'RUA GETULIO VARGAS 116', 'web_name': 'NOVO BARATEIRO'},
	'Supermercado Meira': {'adress': 'AVENIDA JURACY MAGALHAES 426', 'web_name': 'SUPERMERCADOS DALNORDES'},
	'Supermercado Barateiro': {'adress': 'RUA JOSE BONIFACIO 226', 'web_name': 'ITAO'},
	'Supermercado Itao - Centro': {'adress': 'AVENIDA AMELIA AMADO 296', 'web_name': 'ITAO SUPERMERCADOS IMPORTACOES E EXPORTACOES S/A'},
	'Supermercado Itao - S.C': {'adress':'AVENIDA PRINCESA ISABEL 775', 'web_name': 'HIPER ITAO'},

}

LOCALS_ADR_IOS = {

    'Alana Hipermercado': {'adress': 'Rua Santos Dumont 40', 'web_name': 'Alana Hipermercado' },
    'Atacadao': {'adress': 'Rodovia jorge Amado S/n', 'web_name': 'Atacadao'},
   	'Cestao da Economia': {'adress': 'Avenida Lotus 24', 'web_name': 'NAO POSSUI'},
   	'Itao Supermercado': {'adress': 'Avenida Petrobrás S/n', 'web_name': 'Supermercados Itao'},
	'Jaciana Supermercado': {'adress': 'Rua Coronel Pessoa 51', 'web_name': 'Mercado Jaciana'},
	'Gbarbosa': {'adress': 'Avenida Lomando Junior 786', 'web_name': 'GBARBOSA'},
	'Frutaria e Mercadinho Cladudinete': {'adress': 'Avenida Ilhéus/Salobrinho 91', 'web_name': 'Frutaria e Mercadinho Claudinete'},
	'Nenem Supermercados': {'adress': 'Rua Dois de Julho  480', 'web_name': 'Meu Mercado'},
	'Supermercado Meira - Malhado': {'adress': 'Rua Uruguaiana 1187', 'web_name': 'Supermercado Dalnorde'},
	'Supermercado Meira - Centro': {'adress': 'Avenida Coronem Misael Tavares 432', 'web_name': 'Dalnorde Supermercados'},
	'Supermercado Meira - N. Senhora da Vitoria': {'adress': 'Rua São Jorge  64', 'web_name': 'Supermercado Meira'},
	'Supermercado Meira - Vilela': {'adress': 'Avenida Governador Paulo Souto  865', 'web_name': 'Supermercado Meira'},
	'Supermercado Meira - N. Costa': {'adress': 'Rua Jacaranda 250', 'web_name': 'Dalnorde Supermercados'},
	'Supermercado Mangostao': {'adress': 'Avenida Lindolfo Collor 101', 'web_name': 'Supermercado Mangostao LTDA'},

}

PRODUCTS = {'ACUCAR CRISTAL': ['ACUCAR CRISTAL', 'ACUCAR CRISTAL 1KG'],
            'ARROZ PARBOILIZADO': ['ARROZ PARBOILIZADO', 'ARROZ PARBOILIZADO 1KG'],
            'BANANA PRATA': ['BANANA DA PRATA', 'BANANA PRATA', 'BANANA KG'],
            'CAFE MOIDO': ['CAFE 250G', 'CAFE MOIDO'],
            'CHA DENTRO': ['CHA DE DENTRO', 'COXAO MOLE', 'CARNE BOVINA CHA DE DENTRO'],
            'FARINHA DE MANDIOCA': ['FARINHA DE MANDIOCA', 'FARINHA MAND', 'FARINHA MANDIOCA'],
            'FEIJAO CARIOCA': ['FEIJAO CARIOCA'],
            'LEITE LIQUIDO': ['LEITE LIQUIDO'],
            'MANTEIGA 500G': ['MANTEIGA 500G', 'MANTEIGA'],
            'OLEO SOJA': ['OLEO DE SOJA', 'OLEO 900ML'],
            'PAO FRANCES': ['PAO FRANCES', 'PAO FRANCES KG'],
            'TOMATE KG': ['TOMATE KG']}

class Config:

	config_folder = ".config"
	config_file = "config.xlsx"
	file_path = "{}/{}".format(config_folder, config_file)
	config_city = [{"Itabuna": LOCALS_ADR_ITN}, {"Ilhéus": LOCALS_ADR_IOS}]
	products = PRODUCTS
	keywords = None
	city = {}
	estab = {}

	def __init__(self):

		self.create_folder()
		self.create_file()

	def open_xlsx(self):

		os.system("start EXCEL.EXE {}".format(self.file_path))

	def create_folder(self):

		if not os.path.exists(self.config_folder):

			os.makedirs(self.config_folder)
			if os.name == 'nt':
				ret = ctypes.windll.kernel32.SetFileAttributesW(self.config_folder, FILE_ATTRIBUTE_HIDDEN)
				# if ret:
					# print('attribute set to Hidden')
				# else:
					# raise ctypes.WinDLL()

	def create_file(self,):

		if not os.path.exists(self.file_path):

			for city in self.config_city:

				for city_name, city_info in city.items():

					info = city_info
					names = [*info]
					adress = list(map(lambda x: x['adress'], info.values()))
					web = list(map(lambda x: x['web_name'], info.values()))

					try:

						file = load_workbook(self.file_path)
						writer = pd.ExcelWriter(self.file_path, engine='openpyxl')
						writer.book = file
						if city_name in file.sheetnames:
							break

					except:

						writer = pd.ExcelWriter(self.file_path, engine='openpyxl')

					temp_df = pd.DataFrame({"ESTABELECIMENTO" : names, "ENDEREÇO PLATAFORMA": adress, "NOME PLATAFORMA": web})
					temp_df = temp_df.to_excel(writer, sheet_name=city_name,  index=False, startrow=0, startcol=1)
					border = Border(left=Side(border_style='thin', color='FF000000'),
									right=Side(border_style='thin',color='FF000000'),
									top=Side(border_style='thin',color='FF000000'),
									bottom=Side(border_style='thin',color='FF000000'),
									diagonal=Side(border_style='thin', color='FF000000'), diagonal_direction=0,
									outline=Side(border_style='thin',color='FF000000'),
									vertical=Side(border_style='thin',color='FF000000'),
									horizontal=Side(border_style='thin', color='FF000000'))

					workbook = writer.book[city_name]

					worksheet = workbook

					for cell in worksheet['B']:

						cell.border = border
						cell.alignment = Alignment(horizontal='center')

					for cell in worksheet['C']:

						cell.border = border
						cell.alignment = Alignment(horizontal='center')

					for cell in worksheet['D']:

						cell.border = border
						cell.alignment = Alignment(horizontal='center')

					for col in worksheet.columns:
						max_length = 0
						column = col[0].column_letter
						for cell in col:
							try:
								if len(str(cell.value)) > max_length:
									max_length = len(str(cell.value))
							except:
								pass
						adjusted_width = (max_length + 2) * 1.2
						worksheet.column_dimensions[column].width = adjusted_width

					writer.save()

			product_names = []

			for key , value in self.products.items():

				product_names.append([key] * len(value))

			product_names = list(itertools.chain.from_iterable(product_names))
			keywords = self.products.values()
			keywords = list(itertools.chain.from_iterable(keywords))
			sheet_name = 'Produtos'

			try:

				file = load_workbook(self.file_path)
				writer = pd.ExcelWriter(self.file_path, engine='openpyxl')
				writer.book = file
				if sheet_name in file.sheetnames:
					return

			except:

				writer = pd.ExcelWriter(self.file_path, engine='openpyxl')

			temp_df = pd.DataFrame({"PRODUTOS" : product_names, "PALAVRA CHAVE": keywords})
			temp_df = temp_df.to_excel(writer, sheet_name=sheet_name,  index=False, startrow=0, startcol=1)
			border = Border(left=Side(border_style='thin', color='FF000000'),
							right=Side(border_style='thin',color='FF000000'),
							top=Side(border_style='thin',color='FF000000'),
							bottom=Side(border_style='thin',color='FF000000'),
							diagonal=Side(border_style='thin', color='FF000000'), diagonal_direction=0,
							outline=Side(border_style='thin',color='FF000000'),
							vertical=Side(border_style='thin',color='FF000000'),
							horizontal=Side(border_style='thin', color='FF000000'))

			workbook = writer.book[sheet_name]

			worksheet = workbook

			for cell in worksheet['B']:

				cell.border = border
				cell.alignment = Alignment(horizontal='center')

			for cell in worksheet['C']:

				cell.border = border
				cell.alignment = Alignment(horizontal='center')

			for col in worksheet.columns:
				max_length = 0
				column = col[0].column_letter
				for cell in col:
					try:
						if len(str(cell.value)) > max_length:
							max_length = len(str(cell.value))
					except:
						pass
				adjusted_width = (max_length + 2) * 1.2
				worksheet.column_dimensions[column].width = adjusted_width

			writer.save()

	def get_config(self,type):


		if os.path.exists(self.file_path):

			xlsx = pd.ExcelFile(self.file_path)
			file = load_workbook(self.file_path)
			sheet_names = file.sheetnames
			for sheet in sheet_names:

				df = pd.read_excel(xlsx, sheet, skiprows=0, index_col=0)

				if sheet == 'Produtos':

					self.keywords = df.values.tolist()

				else:

					self.city[sheet] = df.values.tolist()
					self.estab[sheet] = df['ESTABELECIMENTO'].tolist()



			if type == 'keywords':

				return self.keywords

			elif type == 'info':

				return self.city

		else:

			self.create_file()
			return self.get_config(type)
